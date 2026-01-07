import shutil
import subprocess
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.reader.excel import load_workbook

from pages.base_page import BasePage
from selenium.webdriver.common.by import By

BASE_URL = "https://rushandball.ru/teams/1417#schedule"

LAST_MATCH_DATA_LOCATOR = (
    "(//div[contains(@class, 'schedule-game')]//div[@class='promo__game-row']"
    "/div[normalize-space(.) != '0']/ancestor::div[contains(@class, 'schedule-game')])"
    "[1]//span[@class='ng-binding']"
)


class MatchParser(BasePage):
    def __init__(self, driver):
        super().__init__(driver)
        self.players = {}
        self.goalkeepers = {}
        self.new_players = {}
        self.new_data = {}

    def open_base_match(self):
        self.open(BASE_URL)
        return self

    def get_last_match_data(self):
        # Ваш XPath с ancestor/contains
        match_data = self.find_elements((
            By.XPATH, LAST_MATCH_DATA_LOCATOR
        ))
        return match_data

    def parse_cska_stats(self):
        """
        Парсит Excel файл ЦСКА и возвращает список игроков с матчами и голами

        Returns:
            Список словарей: [{'name': 'Маркина Полина', 'matches': '212', 'count': '1330'}, ...]
        """

        try:
            workbook = openpyxl.load_workbook(self.base_stat_path)
            sheet = workbook.active  # Лист1

            data = {}
            # Находим строку с заголовками (пропускаем 1-2 строки с описанием)
            for row in range(4, sheet.max_row + 1):  # Начинаем с 3-й строки
                cell_name = sheet.cell(row=row, column=1).value
                cell_matches = sheet.cell(row=row, column=2).value
                cell_goals = sheet.cell(row=row, column=3).value

                # Пропускаем пустые строки и заголовки
                if cell_name and cell_matches and cell_goals:
                    data[cell_name] = {
                        'matches': int(cell_matches),
                        'count': int(cell_goals),
                    }
                elif not self.players:
                    self.players = data
                    data = {}
            self.goalkeepers = data
            print(f"✅ Спарсено {len(self.players)} игроков и {len(self.goalkeepers)} вратарей  из {self.base_stat_path}")

        except Exception as e:
            print(f"❌ Ошибка парсинга Excel: {e}")

    def open_last_match(self):
        """
        Открывает последний матч
        """

        self.click_element((
            By.XPATH, LAST_MATCH_DATA_LOCATOR
        ))
        self.click_element((
            By.XPATH, "(//a[normalize-space(text())='Статистика'])[1]"
        ))

    def add_new_stats(self):
        """
        Получаем статистику по последнему матчу
        """
        players_locator = "//tr/td[2]/a[normalize-space(text())='{}']/ancestor::tr/td[4]"
        goalkeepers_locator = "(//tr/td[2]/a[normalize-space(text())='{}']/ancestor::tr/td[3])[2]"

        players_el = self.find_elements((
            By.XPATH, "//tr/td[2]"
        ))
        cur_locator = players_locator
        players = [player.text for player in players_el[:-1]]
        for player_name in players:
            if player_name:
                player_count = self.find_elements((
                    By.XPATH, cur_locator.format(player_name)
                ))[0].text
                count = int(player_count.split('/')[0])

                if not self.players.get(player_name) and not self.goalkeepers.get(player_name):
                    self.new_players[player_name] = {
                        'matches': 1,
                        'count': 0 if not count else count
                    }

                if player_count and cur_locator == players_locator:
                    self.players[player_name]['matches'] += 1
                    self.players[player_name]['count'] += count

                elif player_count and cur_locator == goalkeepers_locator:
                    if not self.goalkeepers.get(player_name):
                        self.new_players[player_name]['count'] += count

                        self.goalkeepers[player_name] = self.new_players[player_name]
                        self.new_players.pop(player_name)
                    else:
                        self.goalkeepers[player_name]['matches'] += 1
                        self.goalkeepers[player_name]['count'] += count
            else:
                cur_locator = goalkeepers_locator
                print(f"✅ Статистика полевых игроков обновлена")
        for player in list(self.new_players.keys()):
            self.players[player] = self.new_players[player]
        print(f"✅ Статистика вратарей обновлена")

    def clear_and_rewrite_excel(self):
        """
        1. Загружает Excel файл
        2. УДАЛЯЕТ ВСЕ данные с 3-й строки
        3. Записывает НОВЫЕ данные с теми же именами игроков
        4. Сохраняет структуру/форматирование


        """

        try:
            # Загружаем СТРУКТУРУ
            wb = load_workbook(self.base_stat_path)
            ws = wb.active  # Лист1

            # 1. ОЧИЩАЕМ ДАННЫЕ (оставляем заголовки A1:C3)
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=3):
                for cell in row:
                    cell.value = None

            # 2. Записываем новые данные
            # Полевые игроки
            players = list(self.players.keys())
            for i in range(len(players)):
                player_name = players[i]
                ws[f'A{i + 4}'].value = player_name
                ws[f'B{i + 4}'].value = self.players[player_name]['matches']
                ws[f'C{i + 4}'].value = self.players[player_name]['count']

            # Вратари
            goalkeepers = list(self.goalkeepers.keys())
            for i in range(len(goalkeepers)):
                goalkeeper_name = goalkeepers[i]
                ws[f'A{i + 5 + len(players)}'].value = goalkeeper_name
                ws[f'B{i + 5 + len(players)}'].value = self.goalkeepers[goalkeeper_name]['matches']
                ws[f'C{i + 5 + len(players)}'].value = self.goalkeepers[goalkeeper_name]['count']

            wb.save(self.base_stat_path)

            print(f"✅ Очищено и перезаписано: {self.base_stat_path}")

        except Exception as e:
            print(f"❌ Ошибка: {e}")
            return None

    def archive_excel_file(self, new_filename):
        """
        Копирует Excel файл из корня в archive_results с timestamp в названии

        Args:
            new_filename: Базовое имя (по умолчанию - имя исходного файла)

        Returns:
            Path к новому файлу или None при ошибке
        """

        new_name = f"{new_filename}{self.base_stat_path.suffix}"
        destination_file = self.archive_dir / new_name
        source_path = Path(self.base_stat_path).resolve()  # Абсолютный путь
        dest_path = Path(destination_file).resolve()

        # 🔥 ROBOCOPY - ОДИН РАЗ И НАВСЕГДА
        cmd = [
            'robocopy',
            str(source_path.parent),
            str(dest_path.parent),
            source_path.name,
            '/COPY:DAT',  # Data, Attributes, Timestamps
            '/R:0',  # Не повторять при ошибках
            '/W:0'  # Без ожидания
        ]

        result = subprocess.run(cmd, capture_output=True, text=True)

        if result.returncode in [0, 1]:  # robocopy возвращает 1 при успехе
            print(f"✅ Windows копия: {source_path.name} → {dest_path}")
            return dest_path
        else:
            print(f"❌ robocopy ошибка: {result.stderr}")
            return None
